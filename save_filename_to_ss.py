import os
import openpyxl

wb = openpyxl.load_workbook('300.xlsx')

sheet = wb['Sheet1']

path = '/Users/main/Desktop/anki audio'
os.chdir(path)
files = os.listdir(path)

keywords = []

for file in files:
    str = file.replace('.mp3', '')
    str2 = str.replace('(1)', '')
    keywords.append(str2)

os.chdir('/Users/main/pyprojects/forvo')

for i in range(2, 2560):
    old_value = (sheet.cell(row=i, column=1).value)
    new_value = keywords[i-2]
    # print(new_value)
    sheet[f'A{i}'].value = new_value

wb.save('300.xlsx')
print(i, sheet.cell(row=i, column=1).value)
